"""
app.py — 배송 경로 자동 정리 프로그램
디자인: 화이트 베이스 + 블루 포인트 / 라운드 / 클린 모던

주소 검색: 카카오 우편번호 서비스
  → 로컬 HTTP 서버(랜덤 포트) + 시스템 브라우저 방식
  → pywebview / multiprocessing 미사용 (exe 환경 완전 호환)
"""

import http.server
import json
import os
import socket
import subprocess
import sys
import threading
import time
import webbrowser

import customtkinter as ctk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook

BASE = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE)

from core.geocoder  import geocode, reverse_geocode, verify_address
from core.optimizer import build_time_matrix, optimize_route, clear_checkpoint

# ── 디자인 토큰 ───────────────────────────────────────────────────────────────
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

_BG       = "#F7F9FC"
_WHITE    = "#FFFFFF"
_BLUE     = "#2563EB"
_BLUE_LT  = "#EFF6FF"
_BLUE_MID = "#DBEAFE"
_BORDER   = "#E2E8F0"
_TEXT     = "#1E293B"
_SUBTEXT  = "#64748B"
_SUCCESS  = "#16A34A"
_WARN     = "#D97706"
_DANGER   = "#DC2626"
_SHADOW   = "#F1F5F9"


# ─────────────────────────────────────────────────────────────────────────────
# 카카오 우편번호 — 로컬 HTTP 서버 + 시스템 브라우저
# ─────────────────────────────────────────────────────────────────────────────
def _get_free_port() -> int:
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(('', 0))
        return s.getsockname()[1]


def _make_postcode_html(port: int) -> str:
    """
    카카오 우편번호 embed 페이지.
    주소 선택 → localhost:<port>/result POST → 창 자동 닫기 시도.
    """
    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>주소 검색</title>
<style>
  * {{ margin:0; padding:0; box-sizing:border-box; }}
  body {{ font-family: 'Malgun Gothic', sans-serif; background:#F7F9FC; }}
  #header {{
    background:#2563EB; color:#fff;
    padding:14px 20px; font-size:15px; font-weight:700;
    display:flex; align-items:center; gap:8px;
  }}
  #wrap {{ width:100%; height:calc(100vh - 52px); }}
  #done {{
    display:none; position:fixed; inset:0;
    background:#fff;
    flex-direction:column; align-items:center; justify-content:center; gap:14px;
  }}
  .done-icon  {{ font-size:52px; }}
  .done-msg   {{ font-size:18px; font-weight:700; color:#1E293B; }}
  .done-sub   {{ font-size:13px; color:#64748B; }}
  .done-kbd   {{
    background:#F1F5F9; border:1px solid #CBD5E1;
    border-radius:6px; padding:4px 10px;
    font-size:13px; font-family:monospace; color:#475569;
  }}
  .close-btn  {{
    margin-top:4px; padding:10px 32px;
    background:#2563EB; color:#fff; border:none;
    border-radius:8px; font-size:14px; font-weight:600; cursor:pointer;
  }}
  .close-btn:hover {{ background:#1D4ED8; }}
  .countdown  {{ font-size:12px; color:#94A3B8; }}
</style>
<script src="//t1.kakaocdn.net/mapjsapi/bundle/postcode/prod/postcode.v2.js"></script>
</head>
<body>
<div id="header">🔍&nbsp; 주소 검색 — 원하는 주소를 선택하세요</div>
<div id="wrap"></div>
<div id="done">
  <div class="done-icon">✅</div>
  <div class="done-msg">주소가 선택되었습니다</div>
  <div class="done-sub">프로그램에 자동으로 반영됩니다</div>
  <button class="close-btn" onclick="tryClose()">이 창 닫기</button>
  <div class="countdown" id="cd"></div>
  <div class="done-sub">창이 닫히지 않으면 <span class="done-kbd">Ctrl + W</span> 를 누르세요</div>
</div>
<script>
function tryClose() {{
  window.open('', '_self');
  window.close();
}}
new kakao.Postcode({{
  oncomplete: function(data) {{
    var addr = (data.userSelectedType === 'R') ? data.roadAddress : data.jibunAddress;
    var extra = '';
    if (data.userSelectedType === 'R') {{
      if (data.bname !== '' && /[\ub3d9\ub85c\uac00]$/.test(data.bname))
        extra += data.bname;
      if (data.buildingName !== '' && data.apartment === 'Y')
        extra += (extra ? ', ' : '') + data.buildingName;
      if (extra) extra = ' (' + extra + ')';
    }}
    var payload = JSON.stringify({{ address: addr + extra, zonecode: data.zonecode }});
    fetch('http://localhost:{port}/result', {{
      method: 'POST',
      headers: {{'Content-Type': 'application/json'}},
      body: payload
    }}).then(function() {{
      document.getElementById('wrap').style.display = 'none';
      document.getElementById('done').style.display = 'flex';
      // 3초 카운트다운 후 자동 닫기 시도
      var n = 3;
      var cd = document.getElementById('cd');
      var t = setInterval(function() {{
        cd.textContent = n + '초 후 자동으로 닫힙니다...';
        if (n <= 0) {{ clearInterval(t); tryClose(); }}
        n--;
      }}, 1000);
    }}).catch(function(e) {{
      alert('오류가 발생했습니다: ' + e);
    }});
  }},
  width: '100%', height: '100%'
}}).embed(document.getElementById('wrap'));
</script>
</body>
</html>"""


def _find_chromium_exe() -> str | None:
    """Chrome/Edge 실행 파일 경로 탐색. 없으면 None."""
    candidates = [
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


def open_postcode_browser(timeout=180) -> dict:
    """
    로컬 HTTP 서버 + Chrome/Edge 격리 인스턴스로 카카오 우편번호 UI 표시.

    핵심 전략:
    - --user-data-dir=임시폴더 : 기존 Chrome 인스턴스와 완전 격리 → 새 창 보장, 크기 적용 가능
    - 주소 선택 후 Python이 직접 프로세스 종료 → JS window.close() 보안 정책 우회
    - 브라우저 없으면 webbrowser 모듈 fallback (창 제어 불가)
    """
    import tempfile, shutil

    port       = _get_free_port()
    result     = {}
    done       = threading.Event()
    html_bytes = _make_postcode_html(port).encode('utf-8')

    # ── HTTP 서버 ─────────────────────────────────────────────────────────────
    class _Handler(http.server.BaseHTTPRequestHandler):
        def do_GET(self):
            self.send_response(200)
            self.send_header('Content-Type', 'text/html; charset=utf-8')
            self.send_header('Content-Length', str(len(html_bytes)))
            self.end_headers()
            self.wfile.write(html_bytes)

        def do_POST(self):
            try:
                body = self.rfile.read(int(self.headers.get('Content-Length', 0)))
                result.update(json.loads(body))
            except Exception:
                pass
            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Content-Type', 'application/json')
            self.end_headers()
            self.wfile.write(b'{"ok":true}')
            done.set()  # 주소 수신 → 대기 해제

        def do_OPTIONS(self):
            self.send_response(200)
            self.send_header('Access-Control-Allow-Origin', '*')
            self.send_header('Access-Control-Allow-Methods', 'POST, GET, OPTIONS')
            self.send_header('Access-Control-Allow-Headers', 'Content-Type')
            self.end_headers()

        def log_message(self, *_):
            pass

    server = http.server.HTTPServer(('localhost', port), _Handler)
    server.timeout = 1

    def _serve():
        while not done.is_set():
            server.handle_request()
        server.server_close()

    threading.Thread(target=_serve, daemon=True).start()

    # ── 브라우저 실행 ─────────────────────────────────────────────────────────
    exe      = _find_chromium_exe()
    proc     = None
    tmp_dir  = None
    url      = f'http://localhost:{port}'

    if exe:
        # 격리된 임시 프로파일 → 기존 인스턴스와 완전히 분리된 새 창
        tmp_dir = tempfile.mkdtemp(prefix='postcode_')
        flags = [
            f'--user-data-dir={tmp_dir}',  # 격리 인스턴스 (핵심)
            '--new-window',
            '--window-size=500,680',
            '--window-position=300,80',
            '--no-first-run',
            '--no-default-browser-check',
            '--disable-extensions',
            url,
        ]
        try:
            proc = subprocess.Popen(
                [exe] + flags,
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
            )
        except Exception:
            proc = None

    if proc is None:
        # fallback: 기본 브라우저 (창 크기 제어 불가)
        webbrowser.open_new(url)

    # ── 주소 선택 대기 ────────────────────────────────────────────────────────
    done.wait(timeout=timeout)

    # ── 브라우저 프로세스 종료 ────────────────────────────────────────────────
    if proc is not None:
        try:
            proc.terminate()
            proc.wait(timeout=3)
        except Exception:
            try:
                proc.kill()
            except Exception:
                pass

    # 임시 프로파일 폴더 정리 (백그라운드)
    if tmp_dir:
        def _cleanup():
            time.sleep(1)
            shutil.rmtree(tmp_dir, ignore_errors=True)
        threading.Thread(target=_cleanup, daemon=True).start()

    return result


# ─────────────────────────────────────────────────────────────────────────────
# UI 헬퍼
# ─────────────────────────────────────────────────────────────────────────────
def _card(parent, **kw):
    return ctk.CTkFrame(parent, fg_color=_WHITE,
                        corner_radius=16, border_width=1,
                        border_color=_BORDER, **kw)


def _label_sm(parent, text):
    ctk.CTkLabel(parent, text=text,
                 font=ctk.CTkFont(size=11, weight="bold"),
                 text_color=_SUBTEXT).pack(anchor="w", pady=(0, 5))


def _divider(parent):
    ctk.CTkFrame(parent, fg_color=_BORDER, height=1,
                 corner_radius=0).pack(fill="x", pady=8)


# ─────────────────────────────────────────────────────────────────────────────
# 출발지 주소 찾기 팝업
# ─────────────────────────────────────────────────────────────────────────────
class AddressSearchDialog(ctk.CTkToplevel):
    def __init__(self, parent, headers: dict):
        super().__init__(parent)
        self.headers = headers
        self.result  = None
        self.title("출발지 주소 찾기")
        self.geometry("460x240")
        self.resizable(False, False)
        self.configure(fg_color=_BG)
        self.grab_set()
        self._build()

    def _build(self):
        card  = _card(self)
        card.pack(fill="both", expand=True, padx=18, pady=18)
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=20, pady=18)

        ctk.CTkLabel(inner, text="🏢  출발 창고 / 사무실 주소",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=_TEXT).pack(anchor="w", pady=(0, 2))
        ctk.CTkLabel(inner,
                     text="버튼을 누르면 브라우저에서 카카오 주소 검색이 열립니다",
                     font=ctk.CTkFont(size=11), text_color=_SUBTEXT).pack(
            anchor="w", pady=(0, 14))

        self.search_btn = ctk.CTkButton(
            inner, text="🔍   주소 검색 창 열기",
            height=44, corner_radius=10,
            fg_color=_BLUE, hover_color="#1D4ED8",
            font=ctk.CTkFont(size=13, weight="bold"),
            command=self._launch)
        self.search_btn.pack(fill="x", pady=(0, 10))

        self.status = ctk.CTkLabel(inner, text="",
                                   font=ctk.CTkFont(size=11),
                                   text_color=_SUBTEXT, wraplength=380)
        self.status.pack(anchor="w", pady=(0, 10))

        row = ctk.CTkFrame(inner, fg_color="transparent")
        row.pack(fill="x")
        self.ok_btn = ctk.CTkButton(
            row, text="확인", height=38, corner_radius=10,
            fg_color=_BLUE, hover_color="#1D4ED8",
            state="disabled", command=self._confirm)
        self.ok_btn.pack(side="left", expand=True, padx=(0, 6))
        ctk.CTkButton(row, text="닫기", height=38, corner_radius=10,
                      fg_color=_BORDER, hover_color="#CBD5E1",
                      text_color=_TEXT,
                      command=self.destroy).pack(side="left", expand=True)

    def _launch(self):
        self.search_btn.configure(state="disabled",
                                  text="⏳  브라우저에서 주소를 선택해주세요...")
        self.status.configure(text="브라우저에서 주소를 선택하면 자동으로 반영됩니다",
                              text_color=_SUBTEXT)
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        raw = open_postcode_browser()
        if raw.get('address'):
            geo = geocode(raw['address'], self.headers)
            if geo:
                self.result = {'address': raw['address'],
                               'lat': geo['lat'], 'lon': geo['lon']}
                self.after(0, self._on_success)
            else:
                self.after(0, self._on_fail)
        else:
            self.after(0, self._on_cancel)

    def _on_success(self):
        self.status.configure(text=f"✅  {self.result['address']}",
                              text_color=_SUCCESS)
        self.search_btn.configure(state="normal",
                                  text="🔍   주소 검색 창 열기")
        # 주소 선택 완료 → 자동 확인 (사용자가 버튼 누를 필요 없음)
        self.after(400, self._confirm)

    def _on_fail(self):
        self.status.configure(text="⚠️  좌표 변환 실패. 다시 시도해주세요.",
                              text_color=_WARN)
        self.search_btn.configure(state="normal",
                                  text="🔍   주소 검색 창 열기")

    def _on_cancel(self):
        self.status.configure(text="검색을 취소했거나 시간이 초과됐습니다",
                              text_color=_SUBTEXT)
        self.search_btn.configure(state="normal",
                                  text="🔍   주소 검색 창 열기")

    def _confirm(self):
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# 주소 불일치 수정 팝업
# ─────────────────────────────────────────────────────────────────────────────
class AddressFixDialog(ctk.CTkToplevel):
    def __init__(self, parent, headers: dict,
                 name: str, orig: str, rev: str,
                 event: threading.Event, result_holder: dict):
        super().__init__(parent)
        self.headers       = headers
        self.event         = event
        self.result_holder = result_holder
        self._pending      = None
        self.title("주소 확인 필요")
        self.geometry("520x430")
        self.resizable(False, False)
        self.configure(fg_color=_BG)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._skip)
        self._build(name, orig, rev)

    def _build(self, name, orig, rev):
        card  = _card(self)
        card.pack(fill="both", expand=True, padx=18, pady=18)

        # 경고 배너
        top = ctk.CTkFrame(card, fg_color="#FEF2F2",
                           corner_radius=10, height=44)
        top.pack(fill="x", padx=14, pady=(14, 0))
        top.pack_propagate(False)
        ctk.CTkLabel(top, text="⚠️   주소 불일치 감지 — 확인이 필요합니다",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=_DANGER).pack(expand=True)

        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=20, pady=10)

        ctk.CTkLabel(inner, text=f"수령인:  {name}",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color=_TEXT).pack(anchor="w", pady=(0, 8))

        for icon, label, text, bg in [
            ("📋", "엑셀 입력 주소",  orig,                    _BLUE_LT),
            ("🗺️", "지도 확인 주소", rev or "(위치 못 찾음)", "#FEF2F2"),
        ]:
            ctk.CTkLabel(inner, text=f"{icon}  {label}",
                         font=ctk.CTkFont(size=10, weight="bold"),
                         text_color=_SUBTEXT).pack(anchor="w")
            ctk.CTkLabel(inner, text=text, wraplength=460,
                         font=ctk.CTkFont(size=11),
                         fg_color=bg, corner_radius=8,
                         padx=12, pady=7,
                         text_color=_TEXT).pack(fill="x", pady=(2, 8))

        self.search_btn = ctk.CTkButton(
            inner, text="🔍   올바른 주소 검색 (브라우저에서 열림)",
            height=40, corner_radius=10,
            fg_color=_BLUE, hover_color="#1D4ED8",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._launch)
        self.search_btn.pack(fill="x", pady=(4, 6))

        self.verify_lbl = ctk.CTkLabel(inner, text="",
                                        font=ctk.CTkFont(size=11),
                                        text_color=_SUBTEXT, wraplength=460)
        self.verify_lbl.pack(anchor="w", pady=(0, 8))

        row = ctk.CTkFrame(inner, fg_color="transparent")
        row.pack(fill="x")
        self.ok_btn = ctk.CTkButton(
            row, text="✅  이 주소로 수정",
            height=38, corner_radius=10,
            fg_color=_BLUE, hover_color="#1D4ED8",
            state="disabled", command=self._confirm)
        self.ok_btn.pack(side="left", expand=True, padx=(0, 6))
        ctk.CTkButton(row, text="원본 그대로 사용",
                       height=38, corner_radius=10,
                       fg_color=_BORDER, hover_color="#CBD5E1",
                       text_color=_TEXT,
                       command=self._skip).pack(side="left", expand=True)

    def _launch(self):
        self.search_btn.configure(state="disabled",
                                  text="⏳  브라우저에서 주소를 선택해주세요...")
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        raw = open_postcode_browser()
        if raw.get('address'):
            geo = geocode(raw['address'], self.headers)
            if geo:
                rev     = reverse_geocode(geo['lat'], geo['lon'], self.headers)
                verdict = verify_address(raw['address'], rev)
                self._pending = {'address': raw['address'], 'lat': geo['lat'],
                                 'lon': geo['lon'], 'reverse': rev,
                                 'verdict': verdict}
                self.after(0, self._on_success, verdict, raw['address'])
            else:
                self.after(0, self._on_fail)
        else:
            self.after(0, self._on_cancel)

    def _on_success(self, verdict, addr):
        color = _SUCCESS if verdict == '일치' else _WARN
        icon  = "✅" if verdict == '일치' else "⚠️"
        self.verify_lbl.configure(text=f"{icon}  선택된 주소: {addr}",
                                   text_color=color)
        self.ok_btn.configure(state="normal")
        self.search_btn.configure(
            state="normal", text="🔍   올바른 주소 검색 (브라우저에서 열림)")
        # 일치하는 경우 자동 확인
        if verdict == '일치':
            self.after(600, self._confirm)

    def _on_fail(self):
        self.verify_lbl.configure(text="⚠️  좌표 변환 실패", text_color=_WARN)
        self.search_btn.configure(
            state="normal", text="🔍   올바른 주소 검색 (브라우저에서 열림)")

    def _on_cancel(self):
        self.verify_lbl.configure(text="검색 취소됨", text_color=_SUBTEXT)
        self.search_btn.configure(
            state="normal", text="🔍   올바른 주소 검색 (브라우저에서 열림)")

    def _confirm(self):
        if self._pending:
            self.result_holder.update(self._pending)
        self.event.set()
        self.destroy()

    def _skip(self):
        self.result_holder.clear()
        self.event.set()
        self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
# 완료 팝업
# ─────────────────────────────────────────────────────────────────────────────
class DoneDialog(ctk.CTkToplevel):
    def __init__(self, parent, output_path: str, warn_cnt: int):
        super().__init__(parent)
        self.title("작업 완료")
        self.geometry("460x240")
        self.resizable(False, False)
        self.configure(fg_color=_BG)
        self.grab_set()

        card  = _card(self)
        card.pack(fill="both", expand=True, padx=18, pady=18)
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="both", expand=True, padx=20, pady=20)

        badge = ctk.CTkFrame(inner, fg_color="#DCFCE7",
                              corner_radius=20, height=34, width=130)
        badge.pack(pady=(0, 10))
        badge.pack_propagate(False)
        ctk.CTkLabel(badge, text="✅  작업 완료",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=_SUCCESS).pack(expand=True)

        folder = os.path.dirname(output_path)
        fname  = os.path.basename(output_path)

        ctk.CTkLabel(inner, text=fname,
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color=_TEXT).pack(pady=(0, 2))
        ctk.CTkLabel(inner, text=folder, wraplength=400,
                     font=ctk.CTkFont(size=10),
                     text_color=_SUBTEXT).pack(pady=(0, 8))

        if warn_cnt:
            ctk.CTkLabel(inner,
                         text=f"⚠️  주소 미수정 {warn_cnt}건 — '주소검증결과' 열 확인",
                         font=ctk.CTkFont(size=11), text_color=_WARN).pack(
                pady=(0, 6))

        row = ctk.CTkFrame(inner, fg_color="transparent")
        row.pack(fill="x")
        ctk.CTkButton(row, text="📂  저장 폴더 열기",
                      height=40, corner_radius=10,
                      fg_color=_BLUE, hover_color="#1D4ED8",
                      command=lambda: self._open(folder)).pack(
            side="left", expand=True, padx=(0, 6))
        ctk.CTkButton(row, text="닫기", height=40, corner_radius=10,
                      fg_color=_BORDER, hover_color="#CBD5E1",
                      text_color=_TEXT,
                      command=self.destroy).pack(side="left", expand=True)

    @staticmethod
    def _open(path):
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            os.system(f'open "{path}"')
        else:
            os.system(f'xdg-open "{path}"')


# ─────────────────────────────────────────────────────────────────────────────
# 메인 앱
# ─────────────────────────────────────────────────────────────────────────────
class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("배송 경로 자동 정리")
        self.geometry("680x860")
        self.resizable(False, False)
        self.configure(fg_color=_BG)

        self.file_path = ""
        self.start_lat = None
        self.start_lon = None
        self._stop_evt = threading.Event()

        self._build_ui()

    def _build_ui(self):
        # 헤더
        hdr = ctk.CTkFrame(self, fg_color=_WHITE, corner_radius=0, height=68)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        ctk.CTkFrame(self, fg_color=_BORDER, height=1,
                     corner_radius=0).pack(fill="x")

        hdr_inner = ctk.CTkFrame(hdr, fg_color="transparent")
        hdr_inner.place(relx=0.5, rely=0.5, anchor="center")

        icon_bg = ctk.CTkFrame(hdr_inner, fg_color=_BLUE_MID,
                                corner_radius=10, width=38, height=38)
        icon_bg.pack(side="left", padx=(0, 12))
        icon_bg.pack_propagate(False)
        ctk.CTkLabel(icon_bg, text="🚚",
                     font=ctk.CTkFont(size=18)).pack(expand=True)

        txt_col = ctk.CTkFrame(hdr_inner, fg_color="transparent")
        txt_col.pack(side="left")
        ctk.CTkLabel(txt_col, text="배송 경로 자동 정리",
                     font=ctk.CTkFont(size=17, weight="bold"),
                     text_color=_TEXT).pack(anchor="w")
        ctk.CTkLabel(txt_col, text="카카오 API 기반 최적 경로 계산",
                     font=ctk.CTkFont(size=11),
                     text_color=_SUBTEXT).pack(anchor="w")

        # 스크롤 본문
        body = ctk.CTkScrollableFrame(
            self, fg_color=_BG,
            scrollbar_fg_color=_BG,
            scrollbar_button_color=_BORDER,
            scrollbar_button_hover_color="#CBD5E1")
        body.pack(fill="both", expand=True)

        G = {"padx": 20, "pady": 6}

        # ① API 키
        c1 = _card(body); c1.pack(fill="x", **G)
        i1 = ctk.CTkFrame(c1, fg_color="transparent")
        i1.pack(fill="x", padx=20, pady=18)
        _label_sm(i1, "① 카카오 REST API 키")
        self.api_entry = ctk.CTkEntry(
            i1, placeholder_text="카카오 REST API 키를 입력하세요",
            height=42, corner_radius=10, show="*",
            fg_color=_SHADOW, border_color=_BORDER, border_width=1,
            text_color=_TEXT, placeholder_text_color="#94A3B8",
            font=ctk.CTkFont(size=12))
        self.api_entry.pack(fill="x")

        # ② 출발지
        c2 = _card(body); c2.pack(fill="x", **G)
        i2 = ctk.CTkFrame(c2, fg_color="transparent")
        i2.pack(fill="x", padx=20, pady=18)
        _label_sm(i2, "② 출발 창고 / 사무실 주소")
        row2 = ctk.CTkFrame(i2, fg_color="transparent")
        row2.pack(fill="x")
        self.addr_entry = ctk.CTkEntry(
            row2, placeholder_text="[주소 찾기]를 눌러 검색하세요",
            height=42, corner_radius=10, state="disabled",
            fg_color=_SHADOW, border_color=_BORDER, border_width=1,
            text_color=_TEXT, placeholder_text_color="#94A3B8",
            font=ctk.CTkFont(size=12))
        self.addr_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        ctk.CTkButton(row2, text="🔍 주소 찾기",
                      width=110, height=42, corner_radius=10,
                      fg_color=_BLUE, hover_color="#1D4ED8",
                      font=ctk.CTkFont(size=12, weight="bold"),
                      command=self._find_origin).pack(side="left")
        self.addr_ok = ctk.CTkLabel(i2, text="",
                                     font=ctk.CTkFont(size=11),
                                     text_color=_SUCCESS)
        self.addr_ok.pack(anchor="w", pady=(6, 0))

        # ③ 파일
        c3 = _card(body); c3.pack(fill="x", **G)
        i3 = ctk.CTkFrame(c3, fg_color="transparent")
        i3.pack(fill="x", padx=20, pady=18)
        _label_sm(i3, "③ 배송 목록 엑셀 파일 (.xlsx)")
        row3 = ctk.CTkFrame(i3, fg_color="transparent")
        row3.pack(fill="x")
        ctk.CTkButton(row3, text="📁 파일 선택",
                      width=110, height=42, corner_radius=10,
                      fg_color=_SHADOW, hover_color=_BLUE_MID,
                      text_color=_TEXT, border_width=1,
                      border_color=_BORDER,
                      font=ctk.CTkFont(size=12),
                      command=self._pick_file).pack(side="left", padx=(0, 8))
        self.file_lbl = ctk.CTkLabel(
            row3, text="선택된 파일 없음",
            text_color=_SUBTEXT, font=ctk.CTkFont(size=11),
            wraplength=440, anchor="w")
        self.file_lbl.pack(side="left", fill="x", expand=True)

        # 실행 / 중단 버튼
        bf = ctk.CTkFrame(body, fg_color="transparent")
        bf.pack(fill="x", padx=20, pady=10)
        self.run_btn = ctk.CTkButton(
            bf, text="▶   배송 순서 자동 정리 시작",
            height=52, corner_radius=12,
            fg_color=_BLUE, hover_color="#1D4ED8",
            font=ctk.CTkFont(size=14, weight="bold"),
            command=self._start)
        self.run_btn.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.stop_btn = ctk.CTkButton(
            bf, text="⏹", height=52, width=64, corner_radius=12,
            fg_color="#FEF2F2", hover_color="#FEE2E2",
            text_color=_DANGER, border_width=1, border_color="#FECACA",
            font=ctk.CTkFont(size=20), state="disabled",
            command=self._request_stop)
        self.stop_btn.pack(side="left")
        self.refresh_btn = ctk.CTkButton(
            bf, text="🔄", height=52, width=64, corner_radius=12,
            fg_color=_SHADOW, hover_color=_BLUE_MID,
            text_color=_TEXT, border_width=1, border_color=_BORDER,
            font=ctk.CTkFont(size=20), state="disabled",
            command=self._refresh)
        self.refresh_btn.pack(side="left", padx=(8, 0))

        # 진행 상황 카드
        c4 = _card(body); c4.pack(fill="x", **G)
        i4 = ctk.CTkFrame(c4, fg_color="transparent")
        i4.pack(fill="x", padx=20, pady=18)

        sr = ctk.CTkFrame(i4, fg_color="transparent")
        sr.pack(fill="x", pady=(0, 6))
        self.step_lbl = ctk.CTkLabel(sr, text="대기 중",
                                      font=ctk.CTkFont(size=12, weight="bold"),
                                      text_color=_SUBTEXT)
        self.step_lbl.pack(side="left")
        self.pct_lbl = ctk.CTkLabel(sr, text="",
                                     font=ctk.CTkFont(size=12, weight="bold"),
                                     text_color=_BLUE)
        self.pct_lbl.pack(side="right")

        self.bar = ctk.CTkProgressBar(i4, height=6, corner_radius=4,
                                       fg_color=_SHADOW,
                                       progress_color=_BLUE)
        self.bar.pack(fill="x", pady=(0, 16))
        self.bar.set(0)

        _divider(i4)
        _label_sm(i4, "작업 현황")
        self.log = ctk.CTkTextbox(
            i4, height=240,
            fg_color=_SHADOW, border_color=_BORDER, border_width=1,
            corner_radius=10,
            font=ctk.CTkFont(family="맑은 고딕", size=11),
            text_color=_TEXT, state="disabled", wrap="word")
        self.log.pack(fill="both", expand=True)

        ctk.CTkFrame(body, fg_color="transparent", height=16).pack()

    # ── 이벤트 ───────────────────────────────────────────────────────────────
    def _find_origin(self):
        k = self.api_entry.get().strip()
        if not k:
            messagebox.showwarning("API 키 필요", "먼저 카카오 API 키를 입력해주세요.")
            return
        dlg = AddressSearchDialog(self, {"Authorization": f"KakaoAK {k}"})
        self.wait_window(dlg)
        if dlg.result:
            r = dlg.result
            self.start_lat, self.start_lon = r['lat'], r['lon']
            self.addr_entry.configure(state="normal")
            self.addr_entry.delete(0, "end")
            self.addr_entry.insert(0, r['address'])
            self.addr_entry.configure(state="disabled")
            self.addr_ok.configure(text=f"✅  확인된 주소: {r['address']}")

    def _pick_file(self):
        p = filedialog.askopenfilename(
            filetypes=[("Excel 파일", "*.xlsx *.xls")])
        if p:
            self.file_path = p
            self.file_lbl.configure(text=os.path.basename(p), text_color=_TEXT)

    def _start(self):
        if not self.api_entry.get().strip():
            messagebox.showwarning("입력 필요", "API 키를 입력해주세요.")
            return
        if self.start_lat is None:
            messagebox.showwarning("입력 필요", "출발지 주소를 선택해주세요.")
            return
        if not self.file_path:
            messagebox.showwarning("입력 필요", "엑셀 파일을 선택해주세요.")
            return
        self._stop_evt.clear()
        self.run_btn.configure(state="disabled", text="⏳  작업 진행 중...")
        self.stop_btn.configure(state="normal")
        self.refresh_btn.configure(state="disabled")
        self._clear_log()
        self.bar.set(0)
        threading.Thread(target=self._pipeline, daemon=True).start()

    def _request_stop(self):
        if not messagebox.askyesno(
                "중단 확인",
                "작업을 중단하시겠습니까?\n\n저장된 도로 계산 데이터도 초기화됩니다."):
            return
        self._stop_evt.set()
        self.stop_btn.configure(state="disabled", text="⏹")
        self._log("⏹  중단 요청 — 현재 단계 완료 후 종료됩니다...")

    # ── 로그 / 진행 헬퍼 ─────────────────────────────────────────────────────
    def _log(self, m):
        self.after(0, self._write_log, m)

    def _write_log(self, m):
        self.log.configure(state="normal")
        self.log.insert("end", m + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def _clear_log(self):
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")

    def _step(self, txt, pct, color=None):
        self.after(0, self.step_lbl.configure,
                   {"text": txt, "text_color": color or _TEXT})
        self.after(0, self.bar.set, pct)
        self.after(0, self.pct_lbl.configure,
                   {"text": f"{int(pct * 100)}%" if pct > 0 else ""})

    def _reset_btn(self):
        self.after(0, self.run_btn.configure,
                   {"state": "normal",
                    "text": "▶   배송 순서 자동 정리 시작"})
        self.after(0, self.stop_btn.configure,
                   {"state": "disabled", "text": "⏹"})
        self.after(0, self.refresh_btn.configure, {"state": "normal"})

    def _refresh(self):
        self._stop_evt.clear()
        self._clear_log()
        self.bar.set(0)
        self.step_lbl.configure(text="대기 중", text_color=_SUBTEXT)
        self.pct_lbl.configure(text="")
        self.run_btn.configure(state="normal",
                               text="▶   배송 순서 자동 정리 시작")
        self.stop_btn.configure(state="disabled", text="⏹")
        self.refresh_btn.configure(state="disabled")

    def _open_fix(self, headers, name, orig, rev, event, holder):
        AddressFixDialog(self, headers, name, orig, rev, event, holder)

    def _stopped(self) -> bool:
        return self._stop_evt.is_set()

    def _abort(self):
        clear_checkpoint()
        self._log("🗑️  저장된 도로 계산 데이터 초기화 완료")
        self._step("중단됨", 0, _DANGER)
        self._reset_btn()

    # ── 파이프라인 ───────────────────────────────────────────────────────────
    def _pipeline(self):
        key     = self.api_entry.get().strip()
        headers = {"Authorization": f"KakaoAK {key}"}

        try:
            # 1단계
            self._step("1단계 — 엑셀 파일 읽는 중", 0.05)
            self._log("─" * 36); self._log("  1단계   엑셀 파일 읽기")
            self._log("─" * 36)
            try:
                df = pd.read_excel(self.file_path, header=0)
            except Exception as e:
                self._log(f"❌  파일 읽기 실패: {e}")
                self._reset_btn(); return
            addr_col = next((c for c in df.columns if '택배받을 주소' in str(c)), None)
            if addr_col is None:
                self._log("❌  '택배받을 주소' 열이 없습니다.")
                self._reset_btn(); return
            if addr_col != '택배받을 주소':
                self._log(f"ℹ️  '{addr_col}' 열을 주소 열로 사용합니다.")
                df = df.rename(columns={addr_col: '택배받을 주소'})
            df = df.dropna(subset=['택배받을 주소'])
            df['택배받을 주소'] = df['택배받을 주소'].astype(str).str.strip()
            total = len(df)
            self._log(f"✅  {total}건 확인 완료")
            if self._stopped(): self._abort(); return

            # 2단계
            self._step("2단계 — 배송지 위치 확인 중", 0.15)
            self._log(f"\n{'─'*36}"); self._log("  2단계   각 배송지 위치 확인")
            self._log("─" * 36)
            lats, lons, k_addrs = [], [], []
            for i, (_, row) in enumerate(df.iterrows()):
                if self._stopped(): self._abort(); return
                r    = geocode(row['택배받을 주소'], headers)
                name = row.get('이름', '')
                time.sleep(0.15)
                if r:
                    lats.append(r['lat']); lons.append(r['lon'])
                    k_addrs.append(r['kakao_road_addr'])
                    self._log(f"  ✅  ({i+1}/{total})  {name}")
                else:
                    lats.append(None); lons.append(None); k_addrs.append('')
                    self._log(f"  ⚠️   ({i+1}/{total})  {name}  — 위치 못 찾음")
                self._step("2단계 — 배송지 위치 확인 중",
                           0.15 + (i + 1) / total * 0.15)
            df['Latitude']      = lats
            df['Longitude']     = lons
            df['카카오_확인주소'] = k_addrs
            self._log(f"\n✅  2단계 완료 — {sum(1 for v in lats if v)}/{total}건")
            if self._stopped(): self._abort(); return

            # 3단계
            self._step("3단계 — 주소 정확도 검사 중", 0.32)
            self._log(f"\n{'─'*36}"); self._log("  3단계   주소 정확도 검사")
            self._log("─" * 36)
            revs, verdicts, fix_cnt = [], [], 0
            for i, (di, row) in enumerate(df.iterrows()):
                if self._stopped(): self._abort(); return
                lat, lon = row['Latitude'], row['Longitude']
                name     = row.get('이름', f'항목{i+1}')
                if pd.notna(lat) and pd.notna(lon):
                    rev     = reverse_geocode(lat, lon, headers)
                    time.sleep(0.15)
                    verdict = verify_address(row['택배받을 주소'], rev)
                else:
                    rev, verdict = '', '위치없음'
                if verdict not in ('일치', '위치없음', '확인불가'):
                    self._log(f"  ⚠️   ({i+1}/{total})  {name}  — 불일치 → 팝업 확인")
                    self._step("⚠️  불일치 — 팝업에서 주소를 검색해주세요",
                               0.32 + (i + 1) / total * 0.13, _WARN)
                    ev, holder = threading.Event(), {}
                    self.after(0, self._open_fix, headers, name,
                               row['택배받을 주소'], rev, ev, holder)
                    ev.wait()
                    if self._stopped(): self._abort(); return
                    if holder:
                        df.at[di, 'Latitude']      = holder['lat']
                        df.at[di, 'Longitude']     = holder['lon']
                        df.at[di, '카카오_확인주소'] = holder['address']
                        rev     = holder.get('reverse', holder['address'])
                        verdict = holder.get('verdict', '수정됨')
                        fix_cnt += 1
                        self._log(f"     →  수정: {holder['address']}")
                    else:
                        self._log("     →  원본 그대로 사용")
                else:
                    icon = "✅" if verdict == '일치' else "➖"
                    self._log(f"  {icon}  ({i+1}/{total})  {name}  — {verdict}")
                revs.append(rev); verdicts.append(verdict)
                self._step("3단계 — 주소 정확도 검사 중",
                           0.32 + (i + 1) / total * 0.13)
            df['역지오코딩_주소'] = revs; df['주소검증결과'] = verdicts
            warn_cnt = sum(1 for v in verdicts
                           if v not in ('일치','위치없음','확인불가','수정됨'))
            self._log(f"\n✅  3단계 완료"
                      + (f"  — {fix_cnt}건 수정 / ⚠️ {warn_cnt}건 미수정"
                         if fix_cnt or warn_cnt else "  — 모두 정상"))
            if self._stopped(): self._abort(); return

            # 4단계
            self._step("4단계 — 도로 이동 시간 계산 중 (시간 소요)", 0.46)
            self._log(f"\n{'─'*36}"); self._log("  4단계   도로 이동 시간 계산")
            self._log("  배송지 수에 따라 10~15분 소요됩니다"); self._log("─" * 36)
            vdf   = df.dropna(subset=['Latitude', 'Longitude']).copy()
            nodes = [{'id': -1, 'name': '출발지',
                      'lat': self.start_lat, 'lon': self.start_lon}]
            for idx, row in vdf.iterrows():
                nodes.append({'id': idx,
                              'name': row.get('이름', f'배송지{idx}'),
                              'lat': row['Latitude'], 'lon': row['Longitude'],
                              'address': row.get('택배받을 주소', '')})

            def _prog(done, tot):
                if self._stopped(): return
                self._step("4단계 — 도로 이동 시간 계산 중",
                           0.46 + done / tot * 0.24)
                self._log(f"  →  {done} / {tot} 경로 계산 완료")

            clear_checkpoint()
            matrix = build_time_matrix(nodes, headers,
                                       progress_cb=_prog,
                                       stop_event=self._stop_evt)
            if self._stopped(): self._abort(); return
            self._log(f"✅  4단계 완료 — {len(nodes)}개 지점")

            # 5단계
            self._step("5단계 — 최적 배송 순서 계산 중 (최대 3분)", 0.72)
            self._log(f"\n{'─'*36}"); self._log("  5단계   최적 배송 순서 계산 (최대 3분)")
            self._log("─" * 36)
            ordered = optimize_route(nodes, matrix)
            if self._stopped(): self._abort(); return
            if ordered is None:
                self._log("❌  순서 계산 실패"); self._reset_btn(); return
            mapping = {nodes[ni]['id']: step
                       for step, ni in enumerate(ordered, 1)}
            self._log(f"✅  5단계 완료 — {len(ordered)}건 순서 결정")

            # 6단계
            self._step("6단계 — 결과 파일 저장 중", 0.90)
            self._log(f"\n{'─'*36}"); self._log("  6단계   결과 파일 저장")
            self._log("─" * 36)
            out = self._save_xlsx(mapping, df)
            if out is None:
                self._reset_btn(); return
            self._step("✅  모든 작업 완료!", 1.0, _SUCCESS)
            self._log(f"\n🎉  완료!")
            self._log(f"    파일명: {os.path.basename(out)}")
            self._log(f"    위치:   {os.path.dirname(out)}")
            self.after(0, lambda: DoneDialog(self, out, warn_cnt))

        except Exception as e:
            self._log(f"\n❌  오류: {e}")
            self.after(0, lambda: messagebox.showerror("오류", str(e)))
        finally:
            self._reset_btn()

    def _save_xlsx(self, mapping: dict, df: pd.DataFrame):
        try:
            base = os.path.splitext(self.file_path)[0]

            # ── ① 배송순서 열 작성 후 오름차순 정렬된 xlsx 저장 ──────────────
            wb   = load_workbook(self.file_path)
            ws   = wb.active
            hrow = [c.value for c in ws[1]]

            if '배송순서' in hrow:
                col = hrow.index('배송순서') + 1
                self._log("  '배송순서' 열 존재 → 덮어씁니다")
            else:
                ws.insert_cols(1)
                ws.cell(row=1, column=1, value='배송순서')
                col = 1
                self._log("  '배송순서' 열 없음 → 첫 번째 열에 추가")

            # 배송순서 값 입력
            for df_idx, order_val in mapping.items():
                ws.cell(row=df_idx + 2, column=col, value=order_val)

            # 데이터 행 전체를 배송순서 기준 오름차순 정렬
            data_rows = list(ws.iter_rows(min_row=2, values_only=True))
            data_rows.sort(key=lambda r: (r[col - 1] is None, r[col - 1]))
            for row_idx, row_data in enumerate(data_rows, start=2):
                for c_idx, val in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=c_idx, value=val)

            out_xlsx = f"{base}_배송순서완성.xlsx"
            wb.save(out_xlsx)
            self._log(f"✅  xlsx 저장: {os.path.basename(out_xlsx)}")

            # ── ② CSV 저장 (배송순서, 이름, 택배받을 주소, Latitude, Longitude) ──
            csv_cols = ['배송순서', '이름', '택배받을 주소', 'Latitude', 'Longitude']

            df_out = df.copy()
            df_out['배송순서'] = df_out.index.map(mapping)
            df_out = df_out.dropna(subset=['배송순서'])
            df_out['배송순서'] = df_out['배송순서'].astype(int)
            df_out = df_out.sort_values('배송순서').reset_index(drop=True)

            # 없는 열은 빈 값으로
            for c in csv_cols:
                if c not in df_out.columns:
                    df_out[c] = ''

            out_csv = f"{base}_배송순서완성.csv"
            df_out[csv_cols].to_csv(out_csv, index=False, encoding='utf-8-sig')
            self._log(f"✅  csv 저장: {os.path.basename(out_csv)}")

            return out_xlsx
        except Exception as e:
            self._log(f"❌  저장 실패: {e}")
            return None


if __name__ == "__main__":
    App().mainloop()
